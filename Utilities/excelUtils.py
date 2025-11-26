import openpyxl

class ExcelUtils:

    @staticmethod
    def getRowCount(file, sheetName):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        return sheet.max_row

    @staticmethod
    def getColumnCount(file, sheetName):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        return sheet.max_column

    @staticmethod
    def readData(file, sheetName, rowNum, colNum):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        return sheet.cell(row=rowNum, column=colNum).value

    @staticmethod
    def writeData(file, sheetName, rowNum, colNum, data):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        sheet.cell(row=rowNum, column=colNum).value = data
        workbook.save(file)

    @staticmethod
    def getDataAsList(file, sheetName):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]

        total_rows = sheet.max_row
        total_cols = sheet.max_column

        all_data = []

        # start from row 2 (skip header)
        for r in range(2, total_rows + 1):
            row_data = []
            for c in range(1, total_cols + 1):
                row_data.append(sheet.cell(row=r, column=c).value)
            all_data.append(row_data)

        return all_data
